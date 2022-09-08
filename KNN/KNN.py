import pandas as pd
import numpy as np
import openpyxl
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.metrics import mean_absolute_error

#def data_preprocessing(): #Prepare data to fit format below for algorithm


class KNN:  
    
    def __init__(self, name, dataset_dir, dataset_pred):
        self.name = name
        self.data = pd.read_excel(dataset_dir)
        self.pred = pd.read_excel(dataset_pred)
        
    def accuracy_test(self): #returns the accuracy of the KNN algorithm on a particular dataset
        dataset=self.data.drop(['Category'], axis = 1) #Seperate the prediction column
        category=self.data['Category']
        dataset_train, dataset_test, category_train, category_test = train_test_split(dataset, category, test_size=0.20, random_state=13)
        scaler = StandardScaler() #fit the data into the algorithm 
        scaler.fit(dataset_train)
        dataset_train = scaler.transform(dataset_train)
        dataset_test = scaler.transform(dataset_test)
        classifier = KNeighborsClassifier(n_neighbors=3) #Create the model
        classifier.fit(dataset_train, category_train) #Train the model
        return classifier.score(dataset_test, category_test) # return accruacy of a model based on a 20:80 dataset split
    
    def train_model(self, num_neighbours):
        dataset=self.data.drop(['Category'], axis = 1) #Seperate the prediction column
        category=self.data['Category']
        model = KNeighborsClassifier(num_neighbours)
        model.fit(dataset, category)
        return model
    
    def predict(self, model):
        dataset_pred=self.pred.drop(['Category'], axis = 1)
        predictions = model.predict(dataset_pred)
        print(predictions)
        return predictions
    
    def predict_probaility(self, model, prediction_csv):
        dataset_pred=self.pred.drop(['Category'], axis = 1)
        wb = openpyxl.load_workbook(prediction_csv)
        ws = wb['Sheet1']
        labels = ["Blank","Education","Fitness","Gym","Leisure","Sleep","Work"]
        row_count = 2
        column_count = 7
        for i in labels:
            ws.cell(row = 1, column=column_count).value = i
            column_count = column_count + 1
        column_count = 7
        for j in model.predict_proba(dataset_pred):
            for k in j:
                ws.cell(row = row_count, column=column_count).value = k
                column_count = column_count + 1
            row_count = row_count+1
            column_count = 7
        wb.save('predictions_probability.csv')
    
    def find_optimal_n(self): #Need to test this function to see if it runs slow as it retrains a model 40 times
        error = []
        dataset=self.data.drop(['Category'], axis = 1) #Seperate the prediction column
        category=self.data['Category']
        dataset_train, dataset_test, category_train, category_test = train_test_split(dataset, category, test_size=0.20, random_state=13)
        scaler = StandardScaler() #fit the data into the algorithm 
        scaler.fit(dataset_train)
        dataset_train = scaler.transform(dataset_train)
        dataset_test = scaler.transform(dataset_test)
        for i in range(1, 40):
            knn = KNeighborsClassifier(n_neighbors=i)
            knn.fit(dataset_train, category_train)
            pred_i = knn.predict(dataset_test)
            mae = knn.score(dataset_test, category_test)
            error.append(mae)
        return [error.index(np.max(error)), np.amax(error)]


def test():
    dataset = 'C:\\Users\\Calvin\\Documents\\SES_STUDIO\\KNN\\TheoreticalData.xlsx' 
    prediction_dataset = 'C:\\Users\\Calvin\\Documents\\SES_STUDIO\\KNN\\TestMon.xlsx'
    calvinKNN = KNN("Calvin", dataset, prediction_dataset)
    print(calvinKNN.accuracy_test())
    model = calvinKNN.train_model(5)
    #calvinKNN.predict(model)
    #calvinKNN.predict_probaility(model, prediction_dataset)
    print(calvinKNN.find_optimal_n())



test()